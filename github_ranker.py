import os
import time
import requests
import openpyxl
from datetime import datetime, timezone, timedelta
from openpyxl.utils import get_column_letter

# ============================================================
# GITHUB AGE-BASED TOP REPOSITORY RANKING
# ============================================================
#
# Logic:
#
# Daily   = repositories <= 1 day old
# Weekly  = repositories <= 7 days old
# 1 Month = repositories <= 30 days old
# 2 Months = repositories <= 60 days old
# 3 Months = repositories <= 90 days old
# 6 Months = repositories <= 180 days old
# 1 Year   = repositories <= 365 days old
# 2 Years  = repositories <= 730 days old
# 3 Years  = repositories <= 1095 days old
# All Time = all repositories
#
# Within each interval:
#
#       FILTER BY REPOSITORY AGE
#                   ↓
#       SORT BY TOTAL STARS
#                   ↓
#              RANK #1...
#
# ============================================================

API_URL = "https://api.github.com/search/repositories"

EXCEL_FILE = "github_top_repos_by_interval.xlsx"

# Maximum repositories to put into each interval sheet
TOP_N = 500

# ============================================================
# GITHUB TOKEN
# ============================================================

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")

if not GITHUB_TOKEN:
    raise RuntimeError(
        "GITHUB_TOKEN environment variable not found. "
        "Check the GitHub Actions env configuration."
    )

HEADERS = {
    "Accept": "application/vnd.github+json",
    "X-GitHub-Api-Version": "2022-11-28",
    "Authorization": f"Bearer {GITHUB_TOKEN}"
}

# ============================================================
# INTERVAL DEFINITIONS
# ============================================================

INTERVALS = {
    "Daily": 1,
    "Weekly": 7,
    "1 Month": 30,
    "2 Months": 60,
    "3 Months": 90,
    "6 Months": 180,
    "1 Year": 365,
    "2 Years": 730,
    "3 Years": 1095,
}

# ============================================================
# EXCEL HEADERS
# ============================================================

HEADERS_EXCEL = [
    "Rank",
    "Repository",
    "Stars",
    "Age (Days)",
    "Created",
    "Forks",
    "Watchers",
    "Open Issues",
    "Language",
    "Description",
    "Last Updated",
    "Last Push",
    "URL"
]

# ============================================================
# HELPER: GITHUB API REQUEST
# ============================================================

def github_search(params):

    response = requests.get(
        API_URL,
        params=params,
        headers=HEADERS,
        timeout=60
    )

    if response.status_code != 200:

        print("❌ GitHub API request failed")
        print(f"HTTP Status: {response.status_code}")
        print(f"Response: {response.text[:1000]}")

        response.raise_for_status()

    return response.json()


# ============================================================
# HELPER: FETCH ALL RESULTS FOR QUERY
# ============================================================

def fetch_repositories(query, max_pages=10):

    repositories = []

    for page in range(1, max_pages + 1):

        params = {
            "q": query,
            "sort": "stars",
            "order": "desc",
            "per_page": 100,
            "page": page
        }

        print(
            f"   📡 Page {page}/{max_pages}..."
        )

        data = github_search(params)

        items = data.get("items", [])

        if not items:
            break

        repositories.extend(items)

        # GitHub search has a practical result limit.
        total_count = data.get("total_count", 0)

        if len(repositories) >= total_count:
            break

        if len(items) < 100:
            break

        time.sleep(1)

    return repositories


# ============================================================
# START
# ============================================================

print("=" * 80)
print("🚀 GITHUB AGE-BASED TOP REPOSITORY RANKING")
print("=" * 80)

now = datetime.now(timezone.utc)

print(f"🕒 Current UTC time: {now.isoformat()}")

# ============================================================
# FETCH REPOSITORIES
# ============================================================
#
# IMPORTANT:
#
# We don't simply fetch the global Top 500 anymore.
#
# We fetch repositories by creation-date ranges so that
# highly-starred NEW repositories aren't missed.
#
# ============================================================

print()
print("🔎 Fetching repositories by age...")

all_repos = {}

# ------------------------------------------------------------
# Maximum age we need
# ------------------------------------------------------------

max_age_days = max(INTERVALS.values())

oldest_date = (
    now - timedelta(days=max_age_days)
).strftime("%Y-%m-%d")

today_date = now.strftime("%Y-%m-%d")

# Search repositories created within the maximum age window.
#
# GitHub's search result limit means we use several smaller
# age buckets instead of one enormous search.
# ------------------------------------------------------------

SEARCH_BUCKETS = [
    ("0-1d", 1),
    ("1-7d", 7),
    ("7-30d", 30),
    ("30-90d", 90),
    ("90-180d", 180),
    ("180-365d", 365),
    ("365-730d", 730),
    ("730-1095d", 1095),
]

previous_days = 0

for bucket_name, bucket_end_days in SEARCH_BUCKETS:

    start_date = (
        now - timedelta(days=bucket_end_days)
    ).strftime("%Y-%m-%d")

    if previous_days == 0:
        end_date = today_date
    else:
        end_date = (
            now - timedelta(days=previous_days)
        ).strftime("%Y-%m-%d")

    query = (
        f"created:{start_date}..{end_date}"
        " stars:>0"
    )

    print()
    print(
        f"🔹 Bucket {bucket_name}: "
        f"{start_date} → {end_date}"
    )

    try:

        repos = fetch_repositories(
            query,
            max_pages=10
        )

        print(
            f"   ✅ Retrieved {len(repos)} repositories"
        )

        for repo in repos:
            all_repos[repo["id"]] = repo

    except Exception as e:

        print(
            f"   ⚠️ Failed bucket {bucket_name}: {e}"
        )

    previous_days = bucket_end_days

# ============================================================
# ALL-TIME TOP 500
# ============================================================
#
# All Time is handled separately.
# ============================================================

print()
print("🌎 Fetching All-Time Top repositories...")

all_time_repos = fetch_repositories(
    "stars:>0",
    max_pages=5
)

for repo in all_time_repos:
    all_repos[repo["id"]] = repo

print(
    f"✅ Total unique repositories collected: "
    f"{len(all_repos)}"
)

if not all_repos:
    raise RuntimeError(
        "No repositories were returned by GitHub."
    )

# ============================================================
# CALCULATE AGE
# ============================================================

processed_repos = []

for repo in all_repos.values():

    created_string = repo.get("created_at")

    if not created_string:
        continue

    try:

        created_at = datetime.fromisoformat(
            created_string.replace("Z", "+00:00")
        )

        age_seconds = (
            now - created_at
        ).total_seconds()

        age_days = age_seconds / 86400

    except Exception:
        continue

    repo_copy = dict(repo)

    repo_copy["_age_days"] = age_days

    processed_repos.append(repo_copy)

print(
    f"✅ Processed {len(processed_repos)} repositories"
)

# ============================================================
# OPEN EXCEL
# ============================================================

print()
print(f"📂 Opening Excel: {EXCEL_FILE}")

if not os.path.exists(EXCEL_FILE):

    raise FileNotFoundError(
        f"Excel file not found: {EXCEL_FILE}"
    )

wb = openpyxl.load_workbook(EXCEL_FILE)

print(
    "📑 Existing sheets:",
    ", ".join(wb.sheetnames)
)

# ============================================================
# CREATE / UPDATE INTERVAL SHEETS
# ============================================================

for sheet_name, max_days in INTERVALS.items():

    print()
    print("=" * 80)
    print(
        f"📊 {sheet_name.upper()} "
        f"(AGE <= {max_days} DAYS)"
    )
    print("=" * 80)

    # --------------------------------------------------------
    # Filter by age
    # --------------------------------------------------------

    eligible = [
        repo
        for repo in processed_repos
        if repo["_age_days"] <= max_days
    ]

    # --------------------------------------------------------
    # Sort by CURRENT TOTAL STARS
    # --------------------------------------------------------

    eligible.sort(
        key=lambda x: x.get(
            "stargazers_count", 0
        ),
        reverse=True
    )

    # Top N
    eligible = eligible[:TOP_N]

    print(
        f"✅ Eligible repositories: {len(eligible)}"
    )

    # --------------------------------------------------------
    # Create or access sheet
    # --------------------------------------------------------

    if sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        # Clear existing sheet contents
        if ws.max_row > 0:

            ws.delete_rows(
                1,
                ws.max_row
            )

    else:

        ws = wb.create_sheet(
            sheet_name
        )

    # --------------------------------------------------------
    # Headers
    # --------------------------------------------------------

    for col, header in enumerate(
        HEADERS_EXCEL,
        start=1
    ):

        ws.cell(
            row=1,
            column=col,
            value=header
        )

    # --------------------------------------------------------
    # Data
    # --------------------------------------------------------

    for rank, repo in enumerate(
        eligible,
        start=1
    ):

        created_at = repo.get(
            "created_at",
            ""
        )

        if created_at:
            created_date = created_at[:10]
        else:
            created_date = ""

        row = [

            rank,

            repo.get(
                "full_name",
                ""
            ),

            repo.get(
                "stargazers_count",
                0
            ),

            round(
                repo.get(
                    "_age_days",
                    0
                ),
                2
            ),

            created_date,

            repo.get(
                "forks_count",
                0
            ),

            repo.get(
                "watchers_count",
                0
            ),

            repo.get(
                "open_issues_count",
                0
            ),

            repo.get(
                "language"
            ) or "Unknown",

            repo.get(
                "description"
            ) or "",

            repo.get(
                "updated_at",
                ""
            )[:10],

            repo.get(
                "pushed_at",
                ""
            )[:10],

            repo.get(
                "html_url",
                ""
            )
        ]

        ws.append(row)

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 8,
        "B": 45,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 18,
        "J": 60,
        "K": 16,
        "L": 16,
        "M": 60
    }

    for column, width in widths.items():

        ws.column_dimensions[
            column
        ].width = width

    print(
        f"🏆 #1: "
        f"{eligible[0]['full_name'] if eligible else 'N/A'}"
    )

    if eligible:

        print(
            f"⭐ Stars: "
            f"{eligible[0].get('stargazers_count', 0):,}"
        )

# ============================================================
# ALL TIME
# ============================================================

print()
print("=" * 80)
print("🌎 ALL TIME")
print("=" * 80)

if "All Time" in wb.sheetnames:

    ws = wb["All Time"]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

else:

    ws = wb.create_sheet("All Time")

# Headers

for col, header in enumerate(
    HEADERS_EXCEL,
    start=1
):

    ws.cell(
        row=1,
        column=col,
        value=header
    )

# Sort every collected repository by stars

all_time_sorted = sorted(
    processed_repos,
    key=lambda x: x.get(
        "stargazers_count",
        0
    ),
    reverse=True
)

all_time_sorted = all_time_sorted[:TOP_N]

for rank, repo in enumerate(
    all_time_sorted,
    start=1
):

    created_at = repo.get(
        "created_at",
        ""
    )

    row = [

        rank,

        repo.get(
            "full_name",
            ""
        ),

        repo.get(
            "stargazers_count",
            0
        ),

        round(
            repo.get(
                "_age_days",
                0
            ),
            2
        ),

        created_at[:10],

        repo.get(
            "forks_count",
            0
        ),

        repo.get(
            "watchers_count",
            0
        ),

        repo.get(
            "open_issues_count",
            0
        ),

        repo.get(
            "language"
        ) or "Unknown",

        repo.get(
            "description"
        ) or "",

        repo.get(
            "updated_at",
            ""
        )[:10],

        repo.get(
            "pushed_at",
            ""
        )[:10],

        repo.get(
            "html_url",
            ""
        )
    ]

    ws.append(row)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

widths = {
    "A": 8,
    "B": 45,
    "C": 14,
    "D": 14,
    "E": 14,
    "F": 14,
    "G": 14,
    "H": 16,
    "I": 18,
    "J": 60,
    "K": 16,
    "L": 16,
    "M": 60
}

for column, width in widths.items():
    ws.column_dimensions[
        column
    ].width = width

# ============================================================
# SAVE
# ============================================================

print()
print("=" * 80)
print("💾 SAVING EXCEL")
print("=" * 80)

wb.save(EXCEL_FILE)

print()
print("=" * 80)
print("🎉 GITHUB AGE-BASED RANKING UPDATE COMPLETE")
print("=" * 80)

print(
    f"📅 Snapshot: {now.strftime('%Y-%m-%d %H:%M UTC')}"
)

print(
    f"📊 Repositories collected: "
    f"{len(processed_repos)}"
)

print(
    f"📁 Excel: {EXCEL_FILE}"
)

print()
print("📑 Ranking logic:")

for name, days in INTERVALS.items():

    print(
        f"   {name:10} → age <= {days:4} days → "
        f"ranked by total stars"
    )

print(
    "   All Time  → all repositories → "
    "ranked by total stars"
)

print("=" * 80)
